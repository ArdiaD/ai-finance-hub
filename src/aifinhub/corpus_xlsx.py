"""Export the curated corpus to a rich Excel spreadsheet (read-only view).

Unlike the review export (pending papers + decision dropdown), this dumps the
full information for papers of a given status — by default the approved corpus.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

from .config import ROOT, DB_PATH
from .db import DB

console = Console()
OUT_DIR = ROOT / "review"

# (header, width, wrap, getter)
COLUMNS = [
    ("title", 50, True, lambda p: p.title),
    ("authors", 30, True, lambda p: ", ".join(p.authors)),
    ("year", 8, False, lambda p: (p.published or "")[:4]),
    ("themes", 26, True, lambda p: ", ".join(p.themes)),
    ("venue", 24, True, lambda p: p.venue or ""),
    ("source", 12, False, lambda p: p.source),
    ("url", 42, False, lambda p: p.url or ""),
    ("pdf_url", 30, False, lambda p: p.pdf_url or ""),
    ("local_pdf", 28, False, lambda p: os.path.basename(p.pdf_path) if p.pdf_path else ""),
    ("featured", 9, False, lambda p: "yes" if p.featured else ""),
    ("score", 7, False, lambda p: p.score),
    ("abstract", 90, True, lambda p: p.abstract),
    ("fingerprint", 18, False, lambda p: p.fingerprint),
]


def export_corpus(status: str = "approved", path: Optional[str] = None) -> Path:
    db = DB(DB_PATH)
    papers = db.query(status=None if status == "all" else status,
                      order="published DESC, fetched_at DESC")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = Path(path) if path else OUT_DIR / f"corpus_{status}_{datetime.now():%Y-%m-%d}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "papers"
    hfill, hfont = PatternFill("solid", fgColor="1F4E78"), Font(bold=True, color="FFFFFF")
    for c, (head, width, *_rest) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    for r, p in enumerate(papers, start=2):
        for c, (_h, _w, wrap, get) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=get(p))
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(papers) + 1, 1)}"
    wb.save(out)
    console.print(f"[green]Exported {len(papers)} {status} papers → {out}[/green]")
    return out
