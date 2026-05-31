"""Excel review bridge.

SQLite stays the source of truth; Excel is just the human review surface:

  review-export  →  write pending papers to an .xlsx with a yes/no/feature
                    dropdown in the first column (open it on Dropbox, decide).
  review-import  →  read the decisions back and update SQLite.

Decision values: yes = approve, feature = approve + highlight on LinkedIn,
no = reject, blank = leave pending (re-appears in the next export).
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console

from .config import DB_PATH, EXCEL_DIR
from .db import DB
from .fame import FAME_THRESHOLD

console = Console()

REVIEW_DIR = EXCEL_DIR
DECISIONS = {"yes": ("approved", False), "feature": ("approved", True),
             "no": ("rejected", None)}

# (header, attribute, width, wrap)
COLUMNS = [
    ("decision", None, 12, False),     # editable flag: yes / feature = on hub, no = off
    ("title", "title", 50, True),
    ("authors", "_authors", 28, True),
    ("year", "_year", 7, False),
    ("themes", "_themes", 24, True),
    ("fame", "_fame", 6, False),         # yes/no: relevant to the FAME project
    ("fame_score", "fame_score", 9, False),  # % similarity to the FAME summary
    ("venue", "venue", 22, True),
    ("source", "source", 12, False),
    ("url", "url", 38, False),
    ("local_pdf", "_local_pdf", 26, True),
    ("score", "score", 7, False),
    ("abstract", "abstract", 80, True),
    ("fingerprint", "fingerprint", 18, False),  # key — do not edit
]


def _decision_for(paper) -> str:
    """Pre-fill the decision cell: yes/feature = on the hub, no = not on the hub."""
    if paper.status == "approved":
        return "feature" if paper.featured else "yes"
    return "no"  # pending or rejected → not on the hub


def export_review(path: Optional[str] = None) -> Path:
    """Write a dated full-database snapshot: YYYY-MM-DD_hub_db.xlsx.

    Every paper, with the editable `decision` column pre-filled from its current
    state: yes / feature = on the hub, no = not on the hub. This is both the
    review surface and the weekly archival record of the database.
    """
    db = DB(DB_PATH)
    rows = db.query(order="published DESC, score DESC")
    if not rows:
        console.print("[yellow]No papers to export.[/yellow]")

    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    out = Path(path) if path else REVIEW_DIR / f"{datetime.now():%Y-%m-%d}_hub_db.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "hub_db"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for c, (head, *_rest) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=head)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = COLUMNS[c - 1][2]

    # Data rows
    for r, p in enumerate(rows, start=2):
        values = {
            "title": p.title, "_themes": ", ".join(p.themes),
            "_authors": ", ".join(p.authors),
            "_year": (p.published or "")[:4], "venue": p.venue,
            "source": p.source, "score": p.score, "abstract": p.abstract,
            "url": p.url, "fingerprint": p.fingerprint,
            "_local_pdf": os.path.basename(p.pdf_path) if p.pdf_path else "",
            "fame_score": p.fame_score, "fame_note": p.fame_note,
            "_fame": "yes" if (p.fame_score or 0) >= FAME_THRESHOLD else "",
        }
        for c, (_head, attr, _w, wrap) in enumerate(COLUMNS, 1):
            val = _decision_for(p) if attr is None else values.get(attr, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=wrap, vertical="top")

    # yes/no/feature dropdown on the decision column for all data rows
    dv = DataValidation(type="list", formula1='"yes,no,feature"', allow_blank=True)
    dv.error = "Pick yes, no, or feature"
    dv.prompt = "yes = include · feature = include + highlight on LinkedIn · no = reject"
    ws.add_data_validation(dv)
    last = len(rows) + 1
    dv.add(f"A2:A{max(last, 2)}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last, 1)}"
    wb.save(out)

    console.print(
        f"[green]Wrote DB snapshot ({len(rows)} papers) → {out}[/green]\n"
        "Edit the [bold]decision[/bold] column (yes = in hub · feature = in + "
        "highlight · no = not on hub), save, then:\n"
        f"  [bold]python -m aifinhub review-import '{out}'[/bold]"
    )
    return out


def import_review(path: str) -> dict:
    db = DB(DB_PATH)
    wb = load_workbook(path, read_only=True)
    ws = wb["hub_db"] if "hub_db" in wb.sheetnames else (
        wb["review"] if "review" in wb.sheetnames else wb.active)

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip().lower() if h else "" for h in next(rows)]
    try:
        i_dec = header.index("decision")
        i_fp = header.index("fingerprint")
    except ValueError:
        raise SystemExit("Spreadsheet missing 'decision' or 'fingerprint' column.")

    stats = {"approved": 0, "featured": 0, "rejected": 0, "skipped": 0, "missing": 0}
    for row in rows:
        if not row or i_fp >= len(row):
            continue
        fp = str(row[i_fp]).strip() if row[i_fp] else ""
        decision = str(row[i_dec]).strip().lower() if row[i_dec] else ""
        if not fp:
            continue
        if decision not in DECISIONS:
            stats["skipped"] += 1
            continue
        if not db.get(fp):
            stats["missing"] += 1
            continue
        status, featured = DECISIONS[decision]
        # The PDF stays in the library; accepting/rejecting only flips the flag.
        db.set_status(fp, status, featured=featured)
        if decision == "feature":
            stats["featured"] += 1
        elif decision == "yes":
            stats["approved"] += 1
        else:
            stats["rejected"] += 1

    console.print(
        f"[green]Imported decisions:[/green] approved={stats['approved']} "
        f"featured={stats['featured']} rejected={stats['rejected']} "
        f"left-pending={stats['skipped']}"
        + (f" [red]missing-fp={stats['missing']}[/red]" if stats["missing"] else "")
    )
    console.print("Next: [bold]python -m aifinhub build-site[/bold]")
    return stats
